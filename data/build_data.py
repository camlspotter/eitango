import dotenv
dotenv.load_dotenv()

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell

wb = load_workbook('words.xlsx')
sheet : Worksheet = wb[wb.sheetnames[0]]

from pydantic import BaseModel, Field, validator

class Word(BaseModel):
    no : int = Field(description='No.')
    word : str = Field(description='単語')
    klass : str = Field(description='品詞')
    meaning : str = Field(description='意味')
    example : str = Field(description='例文')
    japanese : str = Field(description='和訳')

words = []
for r in sheet.rows:
    [c_no, c_word, c_klass, c_meaning, c_example, c_japanese] = r
    # Skip the first row where c_no.data_type == 's'
    if c_no.data_type == 'n':
        words.append(
            Word(
                no= c_no.value,
                word= c_word.value, 
                klass= c_klass.value, 
                meaning= c_meaning.value, 
                example= c_example.value, 
                japanese= c_japanese.value
            )
        )

from langchain_core.output_parsers import JsonOutputParser
from langchain_core.messages import HumanMessage, SystemMessage, AIMessage
from langchain_openai import ChatOpenAI

llm = ChatOpenAI(model='gpt-4o-mini')

class Question(BaseModel):
    """英単語問題のデータ"""
    question: str = Field(description="解凍すべき単語が `(    )` で伏せられています。")
    answer: str = Field(description="`(    )` を埋めるべき単語です。")
    wrong: list[str] = Field(description="これを選ぶと不正解になる単語 9 つのカンマで区切られたリスト。answer と似た意味の単号は出現してはいけません。")

class Questions(BaseModel):
    """英単語問題集"""
    questions : list[Question] = Field(description='問題集')
    
parser = JsonOutputParser(pydantic_object=Questions)

system_message = '''\
英語の問題を作成したい。入力として、英単語ひとつと、英文をひとつ挙げます。英単語は英文の中に出現しますが、活用や複数形になっているかもしれません。その英単語を伏せて、その英文の和訳を提示し、その英単語を推測する問題です。出力は次のフィールドを持つ JSON です:

'question': 解凍すべき単語が `(    )` で伏せられています。
'answer': `(    )` を埋めるべき単語です。
'wrong': これを選ぶと不正解になる単語 9 つのカンマで区切られたリスト。正解と似た意味の単号は出現してはいけません。

例えば、

[{ 'word': 'ability', 'sentence': 'He has the ability to solve difficult problems.' },
 { 'word': 'access', 'sentence': 'Students have access to these computers.' },
]

という入力に対しては、

{ 'questions' : [
    { 'question': 'He has the (     ) to solve difficult problems.',
      'answer': 'ability',
      'wrong': ['chance', 'time', 'space', 'will', 'motivation', 'courage', 'man', 'way', 'speed' ] },
    { 'question': 'Students have (     ) to these computers.',
      'answer': 'access',
      'wrong': ["entry", "permission", "way", "approach", "advantage", "right", "path", "contact", "opportunity"] }
  ]
}       

という JSON を生成します。 wrong フィールドの単語が answer と似た意味を持っていないことを必ず確認してください。
'''

class WordQuestion(BaseModel):
    word : Word = Field(description='単語情報')
    question : Question = Field(description='問題')
    
def ask(words : list[Word]) -> list[WordQuestion]:
    qs_ = [ f'{{ "word": "{w.word}", "setense": "{w.example}" }}' for w in words ]
    qs = f'{{ "questions": [ {','.join(qs_)} ] }}'

    messages = [
        SystemMessage(content= system_message + parser.get_format_instructions()),
        HumanMessage(content= qs)
    ]
    res = llm.invoke(messages)
    assert isinstance(res.content, str)
    ques = Questions.model_validate_json(res.content)
    assert len(ques.questions) == len(words)

    return [ WordQuestion(word=w, question=q) for (w, q) in zip(words, ques.questions) ]

import json

with open('questions.json', 'w', encoding='utf-8') as oc:
    for i in range(35, 40):
        start = i * 10 + 1
        end = (i + 1) * 10
        print(start, end)
        wqs = ask(words[start-1:end])
        for wq in wqs:
            if wq.word.word != wq.question.answer:
                print(f'Warning: {wq}')
            oc.write(wq.model_dump_json())
            oc.write('\n')
